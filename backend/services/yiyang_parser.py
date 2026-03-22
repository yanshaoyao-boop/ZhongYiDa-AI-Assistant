import openpyxl

from services.parser_utils import extract_searchable_warehouse_codes


def _safe_float(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return None


def parse_yiyang_excel(file_path: str) -> list:
    all_quotes = []
    source_name = file_path.split("\\")[-1].split("/")[-1]

    wb = None
    sheets_data = {}
    try:
        if file_path.lower().endswith(".xls"):
            import xlrd

            wb = xlrd.open_workbook(file_path)
            for sheet_name in wb.sheet_names():
                sheet = wb.sheet_by_name(sheet_name)
                sheets_data[sheet_name] = [sheet.row_values(row_idx) for row_idx in range(sheet.nrows)]
            wb.release_resources()
            wb = None
        else:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheets_data[sheet_name] = [list(row) for row in sheet.iter_rows(values_only=True)]

        if "卡派价格汇总表" in sheets_data:
            all_rows = sheets_data["卡派价格汇总表"]
            for row_idx, row in enumerate(all_rows):
                if row_idx == 0:
                    continue

                channel = str(row[0]).strip() if row[0] is not None else ""
                warehouse_code = str(row[1]).strip().upper() if len(row) > 1 and row[1] is not None else ""
                if not channel or not warehouse_code or "对应渠道" in channel:
                    continue

                time_val = str(row[8]).strip() if len(row) > 8 and row[8] is not None else ""

                prices_yw = {}
                first_yw = _safe_float(row[2]) if len(row) > 2 else None
                second_yw = _safe_float(row[3]) if len(row) > 3 else None
                cbm_yw = _safe_float(row[4]) if len(row) > 4 else None
                if first_yw:
                    prices_yw["12KG+"] = first_yw
                if second_yw:
                    prices_yw["50KG+"] = second_yw
                if cbm_yw:
                    prices_yw["按方包税1CBM+"] = cbm_yw

                if prices_yw:
                    all_quotes.append(
                        {
                            "渠道": f"{channel}(义乌仓)",
                            "仓库代码": warehouse_code,
                            "时效和赔偿约定": "",
                            "价格体系": prices_yw,
                            "起收重量": "12KG+",
                            "宣称时效": time_val,
                            "附加备注": "",
                            "_source": source_name,
                            "_type": "yiyang_flat",
                        }
                    )

                prices_sg = {}
                first_sg = _safe_float(row[5]) if len(row) > 5 else None
                second_sg = _safe_float(row[6]) if len(row) > 6 else None
                cbm_sg = _safe_float(row[7]) if len(row) > 7 else None
                if first_sg:
                    prices_sg["12KG+"] = first_sg
                if second_sg:
                    prices_sg["50KG+"] = second_sg
                if cbm_sg:
                    prices_sg["按方包税1CBM+"] = cbm_sg

                if prices_sg:
                    all_quotes.append(
                        {
                            "渠道": f"{channel}(深圳/广州仓)",
                            "仓库代码": warehouse_code,
                            "时效和赔偿约定": "",
                            "价格体系": prices_sg,
                            "起收重量": "12KG+",
                            "宣称时效": time_val,
                            "附加备注": "",
                            "_source": source_name,
                            "_type": "yiyang_flat",
                        }
                    )

        summary_sheets = [name for name in sheets_data.keys() if "渠道汇总" in name]
        for sheet_name in summary_sheets:
            all_rows = sheets_data[sheet_name]

            anchor_idx = -1
            for row_idx, row in enumerate(all_rows):
                if any(cell and "分区" in str(cell) for cell in row):
                    anchor_idx = row_idx
                    break

            if anchor_idx == -1 or anchor_idx + 1 >= len(all_rows):
                continue

            header = all_rows[anchor_idx]
            sub_header = all_rows[anchor_idx + 1]

            col_channel = -1
            col_region = -1
            col_time = -1
            for col_idx, value in enumerate(header):
                text = str(value).strip() if value else ""
                if "下单渠道" in text or "渠道名称" in text:
                    col_channel = col_idx
                if "分区" in text or "仓库代码" in text:
                    col_region = col_idx
                if "时效" in text:
                    col_time = col_idx

            if col_region == -1:
                continue

            groups = [
                {"name": "义乌仓", "col_start": col_region + 1, "col_end": col_region + 3},
                {"name": "深圳/广州/华南仓", "col_start": col_region + 4, "col_end": col_region + 6},
            ]
            for group in groups:
                group["headers"] = []
                for col_idx in range(group["col_start"], group["col_end"] + 1):
                    if col_idx < len(sub_header) and sub_header[col_idx]:
                        group["headers"].append((col_idx, str(sub_header[col_idx]).strip()))

            current_channel = ""
            for row_idx in range(anchor_idx + 2, len(all_rows)):
                row = all_rows[row_idx]
                if not any(cell for cell in row):
                    continue

                channel_val = (
                    str(row[col_channel]).strip()
                    if col_channel != -1 and col_channel < len(row) and row[col_channel] is not None
                    else ""
                )
                if channel_val:
                    current_channel = channel_val.split("\n")[0].strip()
                if not current_channel:
                    continue

                region_val = str(row[col_region]).strip() if col_region < len(row) and row[col_region] is not None else ""
                if not region_val or "分区" in region_val:
                    continue

                time_val = (
                    str(row[col_time]).strip()
                    if col_time != -1 and col_time < len(row) and row[col_time] is not None
                    else ""
                )
                target_codes = extract_searchable_warehouse_codes(region_val) or [""]

                for group in groups:
                    prices = {}
                    for col_idx, tier_label in group["headers"]:
                        price = _safe_float(row[col_idx]) if col_idx < len(row) else None
                        if price and price > 0:
                            prices[tier_label] = price

                    if not prices:
                        continue

                    for code in target_codes:
                        all_quotes.append(
                            {
                                "渠道": f"{current_channel}({group['name']})",
                                "目的地区": region_val if len(region_val) < 10 else "多目的地",
                                "仓库代码": code,
                                "价格体系": prices,
                                "宣称时效": time_val.split("\n")[0] if time_val else "",
                                "附加备注": f"Sheet: {sheet_name}",
                                "_source": source_name,
                                "_type": "yiyang_region",
                            }
                        )

    except Exception as exc:
        import traceback

        print(f"Error parse_yiyang_excel {file_path}: {exc}")
        traceback.print_exc()
    finally:
        if wb:
            wb.close()

    return all_quotes
