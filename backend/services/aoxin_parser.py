import os

import openpyxl

from services.parser_utils import extract_searchable_warehouse_codes


def parse_aoxin_excel(file_path: str) -> list:
    all_quotes = []
    source_name = os.path.basename(file_path)
    wb = None

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        for sheet_name in wb.sheetnames:
            if any(
                keyword in sheet_name
                for keyword in ["目录", "服务", "赔付", "公告", "须知", "地址", "分区", "交接单", "标准", "发票", "模板"]
            ):
                continue

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            header_row_idx = -1
            for row_idx, row in enumerate(rows[:10]):
                row_text = " ".join(str(cell) for cell in row if cell is not None)
                if "目的港" in row_text and any(keyword in row_text for keyword in ["运费", "CBM", "KG", "代码"]):
                    header_row_idx = row_idx
                    break

            if header_row_idx == -1:
                continue

            headers = rows[header_row_idx]
            dest_col = 0
            wh_code_cols = []
            price_cols = []
            delivery_fee_col = -1
            time_col = -1

            for col_idx, value in enumerate(headers):
                header_text = str(value).strip() if value else ""
                if any(keyword in header_text for keyword in ["交货代码", "仓库代码", "FBA代码", "交货仓"]):
                    wh_code_cols.append(col_idx)
                    prev_header = str(headers[col_idx - 1]).strip() if col_idx > 0 and headers[col_idx - 1] else ""
                    if col_idx > 0 and not prev_header:
                        wh_code_cols.append(col_idx - 1)
                if any(keyword in header_text for keyword in ["1CBM", "3CBM", "5CBM", "10CBM", "20CBM", "运费", "KG"]):
                    price_cols.append((col_idx, header_text))
                if "派送费" in header_text:
                    delivery_fee_col = col_idx
                if any(keyword in header_text for keyword in ["预计时效", "实效", "船期"]):
                    time_col = col_idx

            if not wh_code_cols and dest_col + 1 < len(headers):
                wh_code_cols.append(dest_col + 1)

            notes = ""
            for row in rows[header_row_idx:]:
                first_val = str(row[0]).strip() if row[0] is not None else ""
                if any(keyword in first_val for keyword in ["计费标准", "偏远", "注意", "说明", "包装", "规定", "备注"]):
                    notes += " | ".join(str(cell).strip() for cell in row if cell is not None) + "; "

            current_dest = ""
            for row_idx in range(header_row_idx + 1, len(rows)):
                row = rows[row_idx]
                first_cell = str(row[0]).strip() if row[0] is not None else ""
                if any(keyword in first_cell for keyword in ["计费标准", "偏远", "注意", "包装"]):
                    break

                prices = {}
                for col_idx, header_text in price_cols:
                    value = row[col_idx] if col_idx < len(row) else None
                    try:
                        price = float(value) if value is not None else None
                    except (TypeError, ValueError):
                        price = None
                    if price is not None and price > 0:
                        prices[header_text] = price

                if not prices and not row[0]:
                    continue

                dest = str(row[dest_col]).strip() if row[dest_col] else current_dest
                if dest:
                    current_dest = dest
                if dest in {"目的港", "目的地"}:
                    continue

                raw_codes = []
                for col_idx in wh_code_cols:
                    if col_idx < len(row) and row[col_idx]:
                        raw_codes.append(str(row[col_idx]).strip())

                combined_code_str = "/".join(raw_codes)
                if not combined_code_str and not prices:
                    continue

                cleaned_whs = extract_searchable_warehouse_codes(combined_code_str) or [""]
                delivery_fee = (
                    str(row[delivery_fee_col]).strip()
                    if delivery_fee_col != -1 and delivery_fee_col < len(row) and row[delivery_fee_col]
                    else ""
                )
                time_val = (
                    str(row[time_col]).strip()
                    if time_col != -1 and time_col < len(row) and row[time_col]
                    else ""
                )

                for warehouse_code in cleaned_whs:
                    all_quotes.append(
                        {
                            "渠道": f"澳鑫-{sheet_name}",
                            "目的地区": dest,
                            "仓库代码": warehouse_code,
                            "价格体系": prices if prices else "见详情",
                            "派送费": delivery_fee,
                            "宣称时效": time_val,
                            "附加备注": notes.strip(),
                            "_source": source_name,
                            "_type": "aoxin",
                        }
                    )

    except Exception as exc:
        print(f"Error parse_aoxin_excel: {exc}")
    finally:
        if wb:
            wb.close()

    return all_quotes
