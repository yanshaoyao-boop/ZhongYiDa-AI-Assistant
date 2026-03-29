import os
import re
from typing import Any

import openpyxl


KG_TIER_PATTERN = re.compile(r"^\s*(\d+(?:\.\d+)?)\s*KG\+\s*$", re.IGNORECASE)
BLOCK_BREAK_KEYWORDS = (
    "以上渠道",
    "渠道附加费",
    "产品附加费",
    "派送附加费",
    "赔偿说明",
    "渠道注意事项",
    "特别备注",
)


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def _safe_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).strip().replace(",", "")
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except (TypeError, ValueError):
        return None


def _extract_region_label(raw_value: str) -> str:
    text = _normalize_text(raw_value)
    if not text:
        return ""

    for region in ("美西", "美中", "美东"):
        if region in text:
            return region

    text = re.sub(r"[\(（].*?[\)）]", "", text).strip()
    return text


def _is_price_header_row(row_texts: list[str]) -> bool:
    has_channel_col = any(cell in {"销售渠道", "渠道"} for cell in row_texts)
    has_kg_tier = any(KG_TIER_PATTERN.match(cell or "") for cell in row_texts)
    return has_channel_col and has_kg_tier


def _build_header_map(row_texts: list[str]) -> tuple[int, int, int, int, list[tuple[int, str]]]:
    channel_col = -1
    region_col = -1
    time_col = -1
    notes_col = -1
    price_cols: list[tuple[int, str]] = []

    for idx, text in enumerate(row_texts):
        if text in {"销售渠道", "渠道"}:
            channel_col = idx
        if ("分区" in text and "邮编" in text) or text == "分区 邮编":
            region_col = idx
        if "提取时效" in text or "参考时效" in text:
            time_col = idx
        if text == "备注":
            notes_col = idx

        tier_match = KG_TIER_PATTERN.match(text)
        if tier_match:
            tier = f"{tier_match.group(1)}KG+"
            price_cols.append((idx, tier))

    return channel_col, region_col, time_col, notes_col, price_cols


def _infer_destination(sheet_name: str, region_text: str) -> str:
    region = _extract_region_label(region_text)
    if region:
        return region
    if "英国" in sheet_name:
        return "英国"
    if "美国" in sheet_name:
        return "美国"
    return sheet_name


def _is_block_break_row(row_texts: list[str]) -> bool:
    first_non_empty = next((cell for cell in row_texts if cell), "")
    if not first_non_empty:
        return False
    return any(keyword in first_non_empty for keyword in BLOCK_BREAK_KEYWORDS)


def parse_tianhang_excel(file_path: str) -> list[dict]:
    all_quotes: list[dict] = []
    source_name = os.path.basename(file_path)
    wb = None

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        for sheet_name in wb.sheetnames:
            if "空派" not in sheet_name:
                continue
            if "空卡" in sheet_name or "目录" in sheet_name:
                continue

            ws = wb[sheet_name]
            rows = [list(row) for row in ws.iter_rows(values_only=True)]
            row_idx = 0

            while row_idx < len(rows):
                row_texts = [_normalize_text(cell) for cell in rows[row_idx]]
                if not _is_price_header_row(row_texts):
                    row_idx += 1
                    continue

                channel_col, region_col, time_col, notes_col, price_cols = _build_header_map(row_texts)
                if channel_col == -1 or not price_cols:
                    row_idx += 1
                    continue

                current_channel = ""
                row_idx += 1
                while row_idx < len(rows):
                    data_row = rows[row_idx]
                    data_texts = [_normalize_text(cell) for cell in data_row]

                    if _is_price_header_row(data_texts):
                        break

                    if _is_block_break_row(data_texts):
                        break

                    if not any(data_texts):
                        row_idx += 1
                        continue

                    channel_value = data_texts[channel_col] if channel_col < len(data_texts) else ""
                    if channel_value and channel_value not in {"销售渠道", "渠道"}:
                        current_channel = channel_value

                    if not current_channel:
                        row_idx += 1
                        continue

                    prices: dict[str, float] = {}
                    for col_idx, tier_label in price_cols:
                        if col_idx >= len(data_row):
                            continue
                        price = _safe_float(data_row[col_idx])
                        if price is None or price <= 0:
                            continue
                        prices[tier_label] = price

                    if not prices:
                        row_idx += 1
                        continue

                    region_value = data_texts[region_col] if region_col != -1 and region_col < len(data_texts) else ""
                    destination = _infer_destination(sheet_name, region_value)
                    time_value = data_texts[time_col] if time_col != -1 and time_col < len(data_texts) else ""
                    notes_value = data_texts[notes_col] if notes_col != -1 and notes_col < len(data_texts) else ""
                    start_weight = min(prices.keys(), key=lambda tier: float(tier.replace("KG+", "")))

                    all_quotes.append(
                        {
                            "渠道": f"天航-{sheet_name}-{current_channel}",
                            "目的地区": destination,
                            "仓库代码": "",
                            "时效和赔偿约定": "",
                            "价格体系": prices,
                            "起收重量": start_weight,
                            "宣称时效": time_value,
                            "附加备注": notes_value,
                            "_source": source_name,
                            "_type": "tianhang_air",
                        }
                    )

                    row_idx += 1

    except Exception as exc:
        print(f"Error parse_tianhang_excel: {exc}")
    finally:
        if wb:
            wb.close()

    return all_quotes
