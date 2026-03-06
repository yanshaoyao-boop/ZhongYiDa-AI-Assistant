import openpyxl
import os
import re

def parse_shangyi_excel(file_path: str) -> list:
    """
    专门为“商壹”设计的报价解析补丁（针对东南亚航线）。
    特点：
    1. 识别新加坡、马来西亚、泰国、菲律宾等东南亚航线。
    2. 处理空运（首重/续重/阶梯KG）和海运（CBM阶梯）两种模式。
    3. 提取渠道描述和备注。
    """
    all_quotes = []
    wb = None
    source_name = os.path.basename(file_path)

    try:
        # 强制 data_only=True 获取计算后的数值
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        for sheet_name in wb.sheetnames:
            # 过滤非报价Sheet
            if any(kw in sheet_name for kw in ["联系", "目录", "须知", "地址", "分区", "模板", "交接", "包装", "Sheet1"]):
                continue

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows: continue

            # 寻找表头行 (通常在 L4-L6)
            header_row_idx = -1
            for idx, row in enumerate(rows[:10]):
                row_str = " ".join([str(c) for c in row if c is not None])
                if any(kw in row_str for kw in ["渠道代码", "交货代码", "代码"]):
                    header_row_idx = idx
                    break
            
            if header_row_idx == -1: continue

            headers = rows[header_row_idx]
            
            # 确定关键列索引
            goods_type_col = -1
            code_col = -1
            country_col = -1
            price_cols = []
            notes_col = -1
            time_col = -1
            vol_ratio_col = -1
            
            for col_idx, h in enumerate(headers):
                h_str = str(h).strip() if h else ""
                if "类型" in h_str or "品名" in h_str:
                    goods_type_col = col_idx
                if any(kw in h_str for kw in ["渠道代码", "交货代码", "代码"]):
                    code_col = col_idx
                if "国家" in h_str or "目的" in h_str:
                    country_col = col_idx
                if any(kw in h_str for kw in ["0.5KG", "KG+", "CBM", "整柜"]):
                    price_cols.append((col_idx, h_str))
                if "备注" in h_str:
                    notes_col = col_idx
                if any(kw in h_str for kw in ["时效", "工作天"]):
                    time_col = col_idx
                if "体积" in h_str or "比重" in h_str:
                    vol_ratio_col = col_idx

            # 如果没找到类型列，可能是复合表头，尝试找附近的
            if goods_type_col == -1: goods_type_col = 1
            if country_col == -1: country_col = 1 # 默认国家也可能在前面的列

            # 提取数据行
            # 商壹的数据行通常紧跟表头，但有些行可能为空或包含长组名（如“马尼拉”、“西马”等）
            curr_country = ""
            for r_idx in range(header_row_idx + 1, len(rows)):
                row = rows[r_idx]
                
                # 检查是否进入了底部的备注区 (通常通过第一列含有特殊字符或长文本判断)
                first_cell = str(row[0]).strip() if row[0] is not None else ""
                if len(first_cell) > 30 or any(kw in first_cell for kw in ["说明", "标准", "注意", "包装"]):
                    break
                
                # 特殊逻辑：检测是否是大区组名（比如“西马普货经济”在第一列）
                row_full_str = " ".join([str(c) for c in row if c is not None])
                if not any(row[code_col:]): # 后面全是空的，可能是组名
                    if row[goods_type_col]:
                        curr_country = str(row[goods_type_col]).strip()
                    continue

                code = str(row[code_col]).strip() if code_col < len(row) and row[code_col] else ""
                if not code: continue

                # 提取价格
                prices = {}
                for c_idx, h_name in price_cols:
                    val = row[c_idx] if c_idx < len(row) else None
                    if val is not None:
                        try:
                            # 处理“单询”、“——”、“-”等占位符
                            sval = str(val).strip()
                            if any(k in sval for k in ["单询", "——", "-"]):
                                continue
                            fval = float(val)
                            if fval > 0:
                                prices[h_name] = fval
                        except: pass
                
                if not prices: continue

                # 获取商品描述
                goods_desc = str(row[goods_type_col]).strip() if row[goods_type_col] else ""
                country = str(row[country_col]).strip() if country_col != goods_type_col and row[country_col] else curr_country
                
                notes = str(row[notes_col]).strip() if notes_col != -1 and notes_col < len(row) and row[notes_col] else ""
                time = str(row[time_col]).strip() if time_col != -1 and time_col < len(row) and row[time_col] else ""
                vol = str(row[vol_ratio_col]).strip() if vol_ratio_col != -1 and vol_ratio_col < len(row) and row[vol_ratio_col] else ""

                all_quotes.append({
                    "渠道": f"商壹-{sheet_name}",
                    "目的地区": f"{country} {goods_desc}".strip(),
                    "仓库代码": code,
                    "价格体系": prices,
                    "计费说明": f"体积:{vol}; 备注:{notes}",
                    "宣称时效": time,
                    "_source": source_name
                })

    except Exception as e:
        print(f"Error parse_shangyi_excel: {e}")
    finally:
        if wb: wb.close()

    return all_quotes
