import os
import re
from typing import Any

import openpyxl

from services.parser_utils import extract_searchable_warehouse_codes


HEADER_KEYWORDS = (
    "国家",
    "地区",
    "目的地",
    "转运",
    "地址",
    "FBA",
    "邮编",
    "仓库",
    "分区",
    "渠道",
    "下单渠道",
    "重量段",
    "备注",
    "时效",
    "派送方式",
    "注意",
)

CHANNEL_HEADER_KEYWORDS = ("渠道", "服务", "运输方式", "下单渠道")
DEST_HEADER_KEYWORDS = ("国家", "地区", "目的地", "转运", "地址", "FBA", "邮编", "仓库", "分区")
TIME_HEADER_KEYWORDS = ("时效", "提取", "开船", "工作日")
NOTE_HEADER_KEYWORDS = ("备注", "注意", "说明", "派送方式")
STOP_SHEET_KEYWORDS = (
    "报价主页",
    "说明",
    "附加费",
    "声明",
    "收件",
    "模板",
    "偏远",
    "操作费",
    "仓库操作",
)

TIER_RE = re.compile(r"(\d+(?:\.\d+)?)\s*(KG\+?|CBM\+?)", re.IGNORECASE)
TIER_RANGE_KG_RE = re.compile(r"(\d+(?:\.\d+)?)\s*[-~]\s*\d+(?:\.\d+)?\s*KG", re.IGNORECASE)
TIER_FALLBACK_RE = re.compile(r"(\d+(?:\.\d+)?)\s*(KG|CBM)", re.IGNORECASE)
WAREHOUSE_RANGE_RE = re.compile(r"([A-Z]{3,5})(\d+)\s*[-~]\s*([A-Z]{0,5})(\d+)")


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def _safe_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace(",", "")
    if not text:
        return None

    numeric = re.search(r"-?\d+(?:\.\d+)?", text)
    if not numeric:
        return None

    try:
        return float(numeric.group(0))
    except ValueError:
        return None


def _is_stop_sheet(sheet_name: str) -> bool:
    return any(keyword in sheet_name for keyword in STOP_SHEET_KEYWORDS)


def _extract_tier_label(text: str) -> str | None:
    raw = _normalize_text(text)
    if not raw:
        return None

    match = TIER_RE.search(raw)
    if match:
        number = match.group(1)
        unit = match.group(2).upper()
        if not unit.endswith("+"):
            unit += "+"
        return f"{number}{unit}"

    range_match = TIER_RANGE_KG_RE.search(raw)
    if range_match:
        return f"{range_match.group(1)}KG+"

    fallback = TIER_FALLBACK_RE.search(raw)
    if fallback:
        return f"{fallback.group(1)}{fallback.group(2).upper()}+"

    return None


def _contains_header_keyword(text: str, keywords: tuple[str, ...] = HEADER_KEYWORDS) -> bool:
    return any(keyword in text for keyword in keywords)


def _looks_like_header_text(text: str) -> bool:
    raw = _normalize_text(text)
    if not raw:
        return False
    return _contains_header_keyword(raw)


def _looks_like_group_label(text: str) -> bool:
    raw = _normalize_text(text)
    if not raw:
        return False
    if _extract_tier_label(raw):
        return False
    if _safe_float(raw) is not None:
        return False
    if _contains_header_keyword(raw):
        return False
    return len(raw) <= 50


def _extract_tier_columns(row_texts: list[str]) -> list[tuple[int, str]]:
    tier_columns: list[tuple[int, str]] = []
    for idx, cell in enumerate(row_texts):
        tier_label = _extract_tier_label(cell)
        if tier_label:
            tier_columns.append((idx, tier_label))
    return tier_columns


def _find_col_by_keywords(row_texts: list[str], keywords: tuple[str, ...]) -> int:
    for idx, text in enumerate(row_texts):
        if text and any(keyword in text for keyword in keywords):
            return idx
    return -1


def _detect_destination_col(header_texts: list[str], prev_texts: list[str], first_tier_col: int) -> int:
    destination_col = _find_col_by_keywords(header_texts, DEST_HEADER_KEYWORDS)
    if destination_col != -1:
        return destination_col

    destination_col = _find_col_by_keywords(prev_texts, DEST_HEADER_KEYWORDS)
    if destination_col != -1:
        return destination_col

    fallback_col = first_tier_col - 1
    return fallback_col if fallback_col >= 0 else 0


def _detect_channel_col(header_texts: list[str], prev_texts: list[str]) -> int:
    channel_col = _find_col_by_keywords(header_texts, CHANNEL_HEADER_KEYWORDS)
    if channel_col != -1:
        return channel_col

    channel_col = _find_col_by_keywords(prev_texts, CHANNEL_HEADER_KEYWORDS)
    if channel_col != -1:
        return channel_col

    return 0


def _detect_note_or_time_col(header_texts: list[str], prev_texts: list[str], keywords: tuple[str, ...]) -> int:
    col = _find_col_by_keywords(header_texts, keywords)
    if col != -1:
        return col
    return _find_col_by_keywords(prev_texts, keywords)


def _extract_group_map_from_row(row_texts: list[str], tier_columns: list[tuple[int, str]]) -> dict[int, str]:
    mapping: dict[int, str] = {}
    current_group = ""

    for col_idx, _ in tier_columns:
        cell = row_texts[col_idx] if col_idx < len(row_texts) else ""
        if _looks_like_group_label(cell):
            current_group = cell
        if current_group:
            mapping[col_idx] = current_group

    return mapping


def _pick_initial_group_map(
    prev_texts: list[str],
    next_texts: list[str],
    tier_columns: list[tuple[int, str]],
) -> dict[int, str]:
    fallback: dict[int, str] = {}
    for candidate in (next_texts, prev_texts):
        candidate_map = _extract_group_map_from_row(candidate, tier_columns)
        if not candidate_map:
            continue
        if len(set(candidate_map.values())) >= 2:
            return candidate_map
        if not fallback:
            fallback = candidate_map
    return fallback


def _find_block_label(all_rows: list[list[Any]], header_idx: int, sheet_name: str) -> str:
    for probe_idx in range(header_idx - 1, max(-1, header_idx - 6), -1):
        probe_row = all_rows[probe_idx]
        first_values = [_normalize_text(cell) for cell in probe_row[:3]]
        candidate = next((text for text in first_values if text), "")
        if not candidate:
            continue
        if _looks_like_header_text(candidate):
            continue
        if _extract_tier_label(candidate):
            continue
        return candidate
    return sheet_name


def _compose_destination(
    row_texts: list[str],
    destination_col: int,
    channel_col: int,
    first_tier_col: int,
) -> str:
    destination = ""
    if 0 <= destination_col < len(row_texts):
        value = row_texts[destination_col]
        if value and not _looks_like_header_text(value):
            destination = value

    if not destination:
        for col in range(first_tier_col - 1, -1, -1):
            if col == channel_col:
                continue
            value = row_texts[col] if col < len(row_texts) else ""
            if not value or _looks_like_header_text(value):
                continue
            if _safe_float(value) is not None:
                continue
            destination = value
            break

    if not destination:
        return ""

    zone_col = destination_col - 1
    if 0 <= zone_col < len(row_texts):
        zone = row_texts[zone_col]
        if zone and zone not in destination and not _looks_like_header_text(zone) and len(zone) <= 12:
            if _safe_float(zone) is None:
                destination = f"{zone} {destination}"

    return destination.strip()


def _normalize_channel_parts(parts: list[str]) -> str:
    normalized_parts: list[str] = []
    seen: set[str] = set()
    for part in parts:
        value = _normalize_text(part)
        if not value:
            continue
        if value in seen:
            continue
        seen.add(value)
        normalized_parts.append(value)
    return " - ".join(normalized_parts)


def _expand_warehouse_ranges(text: str) -> list[str]:
    expanded: list[str] = []
    seen: set[str] = set()

    for match in WAREHOUSE_RANGE_RE.finditer(text.upper()):
        prefix1, start_text, prefix2, end_text = match.groups()
        prefix2 = prefix2 or prefix1
        if prefix1 != prefix2:
            continue

        start = int(start_text)
        end = int(end_text)
        if end < start or end - start > 30:
            continue

        for number in range(start, end + 1):
            code = f"{prefix1}{number}"
            if code not in seen:
                seen.add(code)
                expanded.append(code)

    return expanded


def _extract_warehouse_codes(destination: str) -> list[str]:
    merged: list[str] = []
    seen: set[str] = set()
    for code in _expand_warehouse_ranges(destination) + extract_searchable_warehouse_codes(destination):
        code_upper = code.upper()
        if code_upper not in seen:
            seen.add(code_upper)
            merged.append(code_upper)
    return merged


def _min_tier_label(price_system: dict[str, float]) -> str:
    parsed: list[tuple[float, str]] = []
    for label in price_system:
        match = re.search(r"(\d+(?:\.\d+)?)", label)
        if not match:
            continue
        parsed.append((float(match.group(1)), label))
    if not parsed:
        return ""
    parsed.sort(key=lambda item: item[0])
    return parsed[0][1]


def parse_rongda_excel(file_path: str) -> list[dict[str, Any]]:
    all_quotes: list[dict[str, Any]] = []
    source_name = os.path.basename(file_path)
    workbook = None

    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        for sheet_name in workbook.sheetnames:
            if _is_stop_sheet(sheet_name):
                continue

            sheet = workbook[sheet_name]
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            if not rows:
                continue

            header_blocks: list[tuple[int, list[tuple[int, str]]]] = []
            for row_idx, row in enumerate(rows):
                row_texts = [_normalize_text(cell) for cell in row]
                tier_columns = _extract_tier_columns(row_texts)
                if len(tier_columns) >= 2:
                    header_blocks.append((row_idx, tier_columns))

            if not header_blocks:
                continue

            for block_idx, (header_row_idx, tier_columns) in enumerate(header_blocks):
                header_texts = [_normalize_text(cell) for cell in rows[header_row_idx]]
                prev_texts = (
                    [_normalize_text(cell) for cell in rows[header_row_idx - 1]]
                    if header_row_idx > 0
                    else []
                )
                next_texts = (
                    [_normalize_text(cell) for cell in rows[header_row_idx + 1]]
                    if header_row_idx + 1 < len(rows)
                    else []
                )

                first_tier_col = min(col_idx for col_idx, _ in tier_columns)
                destination_col = _detect_destination_col(header_texts, prev_texts, first_tier_col)
                channel_col = _detect_channel_col(header_texts, prev_texts)
                note_col = _detect_note_or_time_col(header_texts, prev_texts, NOTE_HEADER_KEYWORDS)
                time_col = _detect_note_or_time_col(header_texts, prev_texts, TIME_HEADER_KEYWORDS)
                current_group_map = _pick_initial_group_map(prev_texts, next_texts, tier_columns)
                block_label = _find_block_label(rows, header_row_idx, sheet_name)
                current_channel = block_label

                next_header_row_idx = (
                    header_blocks[block_idx + 1][0] if block_idx + 1 < len(header_blocks) else len(rows)
                )

                for data_idx in range(header_row_idx + 1, next_header_row_idx):
                    data_row = rows[data_idx]
                    data_texts = [_normalize_text(cell) for cell in data_row]
                    if not any(data_texts):
                        continue

                    numeric_tier_count = 0
                    textual_tier_count = 0
                    for col_idx, _ in tier_columns:
                        value = data_texts[col_idx] if col_idx < len(data_texts) else ""
                        if _safe_float(value) is not None:
                            numeric_tier_count += 1
                        elif _looks_like_group_label(value):
                            textual_tier_count += 1

                    if numeric_tier_count == 0 and textual_tier_count >= 2:
                        candidate_group_map = _extract_group_map_from_row(data_texts, tier_columns)
                        if candidate_group_map:
                            current_group_map = candidate_group_map
                        continue

                    row_channel_value = data_texts[channel_col] if channel_col < len(data_texts) else ""
                    if (
                        row_channel_value
                        and not _looks_like_header_text(row_channel_value)
                        and _safe_float(row_channel_value) is None
                    ):
                        current_channel = row_channel_value

                    grouped_prices: dict[str, dict[str, float]] = {}
                    for col_idx, tier_label in tier_columns:
                        raw_value = data_row[col_idx] if col_idx < len(data_row) else None
                        numeric_value = _safe_float(raw_value)
                        if numeric_value is None or numeric_value <= 0:
                            continue

                        group_name = _normalize_text(current_group_map.get(col_idx, ""))
                        grouped_prices.setdefault(group_name, {})[tier_label] = numeric_value

                    if not grouped_prices:
                        continue

                    destination = _compose_destination(
                        data_texts,
                        destination_col=destination_col,
                        channel_col=channel_col,
                        first_tier_col=first_tier_col,
                    )
                    if not destination:
                        continue

                    note = data_texts[note_col] if note_col != -1 and note_col < len(data_texts) else ""
                    transit_time = data_texts[time_col] if time_col != -1 and time_col < len(data_texts) else ""
                    warehouse_codes = _extract_warehouse_codes(destination)

                    for group_name, price_system in grouped_prices.items():
                        channel = _normalize_channel_parts(
                            [
                                "荣达",
                                sheet_name,
                                block_label if block_label != sheet_name else "",
                                current_channel if current_channel != block_label else "",
                                group_name,
                            ]
                        )
                        all_quotes.append(
                            {
                                "渠道": channel,
                                "目的地区": destination,
                                "仓库代码": "/".join(warehouse_codes),
                                "价格体系": price_system,
                                "起收重量": _min_tier_label(price_system),
                                "宣称时效": transit_time,
                                "附加备注": note,
                                "_source": source_name,
                                "_type": "rongda",
                            }
                        )

    except Exception as exc:
        print(f"Error parse_rongda_excel: {exc}")
    finally:
        if workbook:
            workbook.close()

    return all_quotes
