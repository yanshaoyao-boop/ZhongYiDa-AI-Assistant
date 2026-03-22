import re

import openpyxl

from services.parser_utils import extract_searchable_warehouse_codes


JINLIAN_WAREHOUSE_KEYWORDS = [
    "义乌",
    "深圳",
    "东莞",
    "广州",
    "厦门",
    "泉州",
    "上海",
    "宁波",
    "合肥",
    "杭州",
    "中山",
]

JINLIAN_BLACKLIST = ["偏远", "目录", "联系人", "船期", "条款", "收费标准", "规则", "附加费", "查询"]


def _is_valid_sheet(name: str) -> bool:
    return not any(keyword in name for keyword in JINLIAN_BLACKLIST)


def _safe_float(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return None


def _extract_warehouse_codes(code_str: str) -> list[str]:
    codes: list[str] = []
    for raw_part in re.split(r"[/,，\s\n]+", str(code_str or "")):
        part = raw_part.strip().upper()
        if not part:
            continue
        part = re.sub(r"[\(（].*?[\)）]", "", part).strip()
        if not part:
            continue

        range_match = re.match(r"^([A-Z]{3,4})(\d+)-([A-Z]{3,4})(\d+)$", part)
        if range_match and range_match.group(1) == range_match.group(3):
            prefix = range_match.group(1)
            start = int(range_match.group(2))
            end = int(range_match.group(4))
            for index in range(start, end + 1):
                codes.append(f"{prefix}{index}")
            continue

        codes.extend(extract_searchable_warehouse_codes(part))

    deduped = []
    seen = set()
    for code in codes:
        if code not in seen:
            seen.add(code)
            deduped.append(code)
    return deduped


def _detect_warehouse_groups(header_row: list, ref_row: list, start_col: int) -> list[dict]:
    groups: list[dict] = []
    current_group = None
    for col_idx in range(start_col, min(start_col + 30, len(header_row))):
        header_value = str(header_row[col_idx]).strip() if col_idx < len(header_row) and header_row[col_idx] is not None else ""
        ref_value = str(ref_row[col_idx]).strip() if col_idx < len(ref_row) and ref_row[col_idx] is not None else ""

        if any(keyword in header_value for keyword in JINLIAN_WAREHOUSE_KEYWORDS) and header_value:
            first_city = header_value.split("/")[0].strip()
            current_group = {"name": f"{first_city}仓", "start_col": col_idx, "headers": []}
            groups.append(current_group)

        if current_group and ref_value:
            if any(char.isdigit() for char in ref_value) or "KG" in ref_value.upper() or "CBM" in ref_value.upper():
                current_group["headers"].append((col_idx, ref_value))

    return groups


def parse_jinlian_global_excel(file_path: str) -> list:
    all_quotes = []
    source_name = file_path.split("\\")[-1].split("/")[-1]
    wb = None

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        for sheet_name in wb.sheetnames:
            if not _is_valid_sheet(sheet_name):
                continue

            ws = wb[sheet_name]
            all_rows = [list(row) for row in ws.iter_rows(values_only=True)]

            current_product = sheet_name.replace("（新）", "").strip()
            i = 0
            while i < len(all_rows):
                row = all_rows[i]

                for cell in row:
                    cell_text = str(cell).strip() if cell is not None else ""
                    if "下单产品" in cell_text or "JLHP-" in cell_text or "JLKP-" in cell_text:
                        current_product = cell_text.split("下单")[0].strip()
                        break

                anchor_col = -1
                for col_idx, value in enumerate(row):
                    cell_text = str(value).strip() if value is not None else ""
                    if "交货仓" in cell_text or ("义乌" in cell_text and anchor_col == -1):
                        anchor_col = col_idx
                        break

                if anchor_col == -1 or i + 1 >= len(all_rows):
                    i += 1
                    continue

                header_row = row
                price_header_row = all_rows[i + 1]

                region_col = -1
                wh_code_col = -1
                for col_idx, value in enumerate(price_header_row):
                    cell_text = str(value).strip() if value is not None else ""
                    if "区域" in cell_text or "城市" in cell_text or "国家" in cell_text or "分区" in cell_text:
                        region_col = col_idx
                    if "仓库代码" in cell_text or "邮编" in cell_text:
                        wh_code_col = col_idx

                if region_col == -1 and wh_code_col == -1:
                    region_col = 1
                    wh_code_col = 2

                warehouse_groups = _detect_warehouse_groups(header_row, price_header_row, anchor_col)
                if not warehouse_groups:
                    i += 1
                    continue

                time_col = -1
                for col_idx, value in enumerate(header_row):
                    cell_text = str(value).strip() if value is not None else ""
                    if "时效" in cell_text or "提取" in cell_text:
                        time_col = col_idx
                if time_col == -1:
                    for col_idx, value in enumerate(price_header_row):
                        cell_text = str(value).strip() if value is not None else ""
                        if "时效" in cell_text or "提取" in cell_text or "POD" in cell_text:
                            time_col = col_idx

                i += 2
                while i < len(all_rows):
                    data_row = all_rows[i]
                    if not any(cell for cell in data_row[:6] if cell is not None):
                        break

                    row_text = "".join(str(cell) for cell in data_row if cell is not None)
                    if any(keyword in row_text for keyword in ["下单产品", "交货仓", "附加费", "特别说明"]):
                        break

                    region_val = (
                        str(data_row[region_col]).strip()
                        if region_col != -1 and region_col < len(data_row) and data_row[region_col] is not None
                        else ""
                    )
                    wh_code_val = (
                        str(data_row[wh_code_col]).strip()
                        if wh_code_col != -1 and wh_code_col < len(data_row) and data_row[wh_code_col] is not None
                        else ""
                    )

                    if not region_val and not wh_code_val:
                        region_val = str(data_row[1]).strip() if len(data_row) > 1 and data_row[1] is not None else ""
                        wh_code_val = str(data_row[2]).strip() if len(data_row) > 2 and data_row[2] is not None else ""
                        if not region_val and not wh_code_val:
                            i += 1
                            continue

                    region_clean = re.sub(r"[\(（].*?[\)）]", "", region_val).strip()
                    code_list = _extract_warehouse_codes(wh_code_val)
                    if not code_list:
                        code_list = extract_searchable_warehouse_codes(region_clean)
                    if not code_list:
                        code_list = [""]

                    time_val = ""
                    if time_col != -1 and time_col < len(data_row) and data_row[time_col]:
                        maybe_time = str(data_row[time_col]).strip()
                        if "时效" not in maybe_time:
                            time_val = maybe_time.split("\n")[0]

                    for group in warehouse_groups:
                        prices = {}
                        for col_idx, tier_label in group["headers"]:
                            price = _safe_float(data_row[col_idx]) if col_idx < len(data_row) else None
                            if price and price > 0:
                                prices[tier_label] = price

                        if not prices:
                            continue

                        for code in code_list:
                            all_quotes.append(
                                {
                                    "渠道": f"{current_product}({group['name']})",
                                    "目的地区": region_clean,
                                    "仓库代码": code,
                                    "时效和赔偿约定": "",
                                    "价格体系": prices,
                                    "起收重量": "",
                                    "宣称时效": time_val,
                                    "附加备注": f"{region_val} {wh_code_val}".replace("\n", "").strip()[:80],
                                    "_source": source_name,
                                    "_type": "jinlian_global",
                                }
                            )
                    i += 1

    except Exception as exc:
        import traceback

        print(f"Error parse_jinlian_global_excel {file_path}: {exc}")
        traceback.print_exc()
    finally:
        if wb:
            wb.close()

    return all_quotes
