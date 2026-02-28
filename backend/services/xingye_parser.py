import re
import openpyxl

BLACKLIST = ["目录", "必读", "条款", "介绍"]

def _is_valid_sheet(name: str) -> bool:
    for b in BLACKLIST:
        if b in name:
            return False
    return True

def _safe_float(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).strip())
    except (ValueError, TypeError):
        return None

def parse_xingye_excel(file_path: str) -> list:
    """
    专门用于解析'星夜'中欧/欧洲专线格式的解析器。
    特点: "转运目的地" 作为目的地/仓库，国家和仓库混排，多产品列排。
    """
    all_quotes = []
    source_name = file_path.split("\\")[-1].split("/")[-1]

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        for sheet_name in wb.sheetnames:
            if not _is_valid_sheet(sheet_name):
                continue

            ws = wb[sheet_name]
            all_rows = [list(row) for row in ws.iter_rows(values_only=True)]

            i = 0
            current_category = sheet_name
            while i < len(all_rows):
                row = all_rows[i]
                
                dest_col = -1
                for col_idx, cell in enumerate(row):
                    val = str(cell).strip() if cell is not None else ""
                    if "转运目的地" in val or "国家" in val or "目的地" in val:
                        dest_col = col_idx
                        break
                        
                if dest_col == -1:
                    i += 1
                    continue
                    
                # Process header
                prices_headers = []
                for col_idx in range(dest_col + 1, len(row)):
                    h = str(row[col_idx]).strip() if row[col_idx] is not None else ""
                    if "KG" in h.upper() or "CBM" in h.upper():
                        prices_headers.append((col_idx, h))
                        
                current_sub_channel = current_category
                
                i += 1
                while i < len(all_rows):
                    data_row = all_rows[i]
                    if not any(data_row):
                        i += 1
                        continue
                        
                    # Check if sub-channel changes (it usually appears in a column before dest_col)
                    if dest_col > 0:
                        chan_val = str(data_row[1]).strip() if 1 < len(data_row) and data_row[1] is not None else ""
                        if chan_val and "渠道" in chan_val or "派" in chan_val:
                            current_sub_channel = chan_val.replace("\n", " ").strip()
                            
                    dest_val = str(data_row[dest_col]).strip() if dest_col < len(data_row) and data_row[dest_col] is not None else ""
                    if "目的地" in dest_val or "需知" in dest_val or "收费标准" in dest_val:
                        break
                        
                    if dest_val:
                        prices = {}
                        for col_idx, h in prices_headers:
                            v = _safe_float(data_row[col_idx]) if col_idx < len(data_row) else None
                            if v and v > 0:
                                prices[h] = v
                                
                        if prices:
                            
                            # split multiple destinations if any
                            dests = [d.strip() for d in re.split(r'[、/,\n]+', dest_val) if d.strip()]
                            for d in dests:
                                all_quotes.append({
                                    "渠道": f"{sheet_name} - {current_sub_channel}",
                                    "目的地区": d,
                                    "仓库代码": d,  # using destination as warehouse code
                                    "时效和赔偿约定": "",
                                    "价格体系": prices,
                                    "起收重量": "",
                                    "宣称时效": "",
                                    "附加备注": "",
                                    "_source": source_name,
                                    "_type": "xingye"
                                })
                    i += 1

    except Exception as e:
        import traceback
        print(f"Error parsing xingye: {e}")
        traceback.print_exc()

    return all_quotes
