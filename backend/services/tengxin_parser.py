import re
import openpyxl

BLACKLIST = ["导航", "快速查找", "目录", "说明", "赔偿", "船期", "退件"]

def _is_valid_sheet(name: str) -> bool:
    for b in BLACKLIST:
        if b in name:
            return False
    return True

def _safe_float(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).strip())
    except (ValueError, TypeError):
        return None

def parse_tengxin_excel(file_path: str) -> list:
    """
    专门用于解析'腾信'格式报价表的解析器。
    按网格和仓库代码锚点查找。
    """
    all_quotes = []
    wb = None
    source_name = file_path.split("\\")[-1].split("/")[-1]

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        for sheet_name in wb.sheetnames:
            if not _is_valid_sheet(sheet_name):
                continue

            ws = wb[sheet_name]
            all_rows = [list(row) for row in ws.iter_rows(values_only=True)]

            i = 0
            current_channel = sheet_name
            while i < len(all_rows):
                row = all_rows[i]
                
                # Check for channel name
                first_val = str(row[0]).strip() if row[0] is not None else ""
                if first_val and len([c for c in row if c is not None]) <= 3:
                     if any(kw in first_val for kw in ["派", "专线", "快线", "普船", "快船"]):
                          current_channel = first_val.split("\n")[0].strip()

                # Find warehouse column
                wh_col = -1
                for col_idx, cell in enumerate(row):
                    val = str(cell).strip() if cell is not None else ""
                    if "仓库" in val or "代码" in val or "FBA" in val:
                        wh_col = col_idx
                        break
                        
                if wh_col == -1 or i + 1 >= len(all_rows):
                    i += 1
                    continue
                    
                # We found a header block
                header_row = row
                price_header_row = all_rows[i+1]
                
                # Try to map prices
                prices_headers = []
                for col_idx in range(wh_col + 1, len(price_header_row)):
                    h = str(price_header_row[col_idx]).strip() if price_header_row[col_idx] is not None else ""
                    # Or maybe the header is on the same row for some tables
                    if not h:
                        h = str(header_row[col_idx]).strip() if header_row[col_idx] is not None else ""
                    
                    if "KG" in h.upper() or "CBM" in h.upper() or "方" in h:
                        prices_headers.append((col_idx, h))
                        
                if not prices_headers:
                    i += 1
                    continue
                    
                # Scan data
                i += 2
                while i < len(all_rows):
                    data_row = all_rows[i]
                    if not any(data_row):
                        break
                        
                    wh_raw = str(data_row[wh_col]).strip() if wh_col < len(data_row) and data_row[wh_col] is not None else ""
                    if "赔偿" in wh_raw or "注意" in wh_raw or "说明" in wh_raw:
                        break
                        
                    if not wh_raw:
                        i += 1
                        continue
                        
                    # Extract codes
                    wh_codes = []
                    for code in re.split(r'[/,，\n ]+', wh_raw):
                        code = code.strip().upper()
                        if re.match(r'^[A-Z]{2,4}\d+[A-Z]?$', code):
                            wh_codes.append(code)
                    
                    if not wh_codes:
                        if len(wh_raw) < 15:
                             wh_codes = [wh_raw.strip()]
                        else:
                             i += 1
                             continue
                             
                    prices = {}
                    for col_idx, h in prices_headers:
                        v = _safe_float(data_row[col_idx]) if col_idx < len(data_row) else None
                        if v and v > 0:
                            prices[h] = v
                            
                    if prices:
                        for c in wh_codes:
                            all_quotes.append({
                                "渠道": current_channel,
                                "仓库代码": c,
                                "时效和赔偿约定": "",
                                "价格体系": prices,
                                "起收重量": "",
                                "宣称时效": "",
                                "附加备注": "",
                                "_source": source_name,
                                "_type": "tengxin"
                            })
                    i += 1
            
    except Exception as e:
        print(f"Error parsing tengxin: {e}")
    finally:
        if wb:
            wb.close()
        
    return all_quotes
