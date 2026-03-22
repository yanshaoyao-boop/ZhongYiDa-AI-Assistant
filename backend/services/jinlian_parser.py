import re

import openpyxl

from services.parser_utils import extract_searchable_warehouse_codes


JINLIAN_WAREHOUSE_KEYWORDS = [
    "义乌",
    "深圳",
    "东莞",
    "广州",
    "厦门",
    "泉州",
    "上海",
    "宁波",
    "合肥",
    "杭州",
    "中山",
]

JINLIAN_REGION_PATTERNS = [
    r"美东",
    r"美中",
    r"美西",
    r"美东南",
    r"美东区",
]

JINLIAN_BLACKLIST = [
    "目录",
    "各公司联系人",
    "船期表",
    "理赔条款",
    "偏远邮编",
    "海外仓",
    "承运规则",
    "发票模板",
    "偏远地区",
    "定时达",
    "返点销",
]


def _is_valid_sheet(name: str) -> bool:
    return not any(keyword in name for keyword in JINLIAN_BLACKLIST)


def _safe_float(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return None


def _is_region_row(text: str) -> bool:
    return any(re.search(pattern, text) for pattern in JINLIAN_REGION_PATTERNS)


def _detect_warehouse_groups(header_row: list, ref_row: list, start_col: int) -> list[dict]:
    groups: list[dict] = []
    current_group = None
    for col_idx in range(start_col, min(start_col + 30, len(header_row))):
        header_value = str(header_row[col_idx]).strip() if header_row[col_idx] is not None else ""
        ref_value = str(ref_row[col_idx]).strip() if col_idx < len(ref_row) and ref_row[col_idx] is not None else ""

        if any(keyword in header_value for keyword in JINLIAN_WAREHOUSE_KEYWORDS) and header_value:
            current_group = {
                "name": f"{header_value.split('/')[0].strip()}仓",
                "start_col": col_idx,
                "headers": [],
            }
            groups.append(current_group)

        if not current_group and any(keyword in (header_value + ref_value).upper() for keyword in ["KG", "CBM"]):
            current_group = {"name": "通用", "start_col": col_idx, "headers": []}
            groups.append(current_group)

        found_tier = ""
        if "KG" in header_value.upper() or "CBM" in header_value.upper():
            found_tier = header_value
        elif "KG" in ref_value.upper() or "CBM" in ref_value.upper():
            found_tier = ref_value

        if current_group and found_tier:
            current_group["headers"].append((col_idx, found_tier))

    return groups


def parse_jinlian_excel(file_path: str) -> list:
    all_quotes = []
    source_name = file_path.split("\\")[-1].split("/")[-1]
    wb = None

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        for sheet_name in wb.sheetnames:
            if not _is_valid_sheet(sheet_name):
                continue

            ws = wb[sheet_name]
            all_rows = [list(row) for row in ws.iter_rows(values_only=True)]

            current_channel = sheet_name
            i = 0
            while i < len(all_rows):
                row = all_rows[i]

                anchor_col = -1
                region_col = -1
                is_warehouse_mode = False
                for col_idx, value in enumerate(row):
                    cell_text = str(value).strip() if value is not None else ""
                    if "交货仓" in cell_text or "分区" in cell_text or "仓库代码" in cell_text:
                        anchor_col = col_idx
                        if "仓库代码" in cell_text:
                            is_warehouse_mode = True
                    if "分区" in cell_text or "区域" in cell_text or "仓库" in cell_text:
                        region_col = col_idx

                if anchor_col == -1:
                    non_empty = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                    if (
                        len(non_empty) == 1
                        and len(non_empty[0]) > 8
                        and any(keyword in non_empty[0] for keyword in ["快递派", "卡派", "普船", "快船", "海派", "限时达"])
                    ):
                        current_channel = non_empty[0].split("下单")[0].strip()
                    i += 1
                    continue

                if i + 1 >= len(all_rows):
                    i += 1
                    continue

                price_header_row = all_rows[i + 1]
                if not is_warehouse_mode and anchor_col < len(price_header_row):
                    anchor_hint = str(price_header_row[anchor_col]).strip() if price_header_row[anchor_col] is not None else ""
                    if "仓库" in anchor_hint or "代码" in anchor_hint or "FBA" in anchor_hint:
                        is_warehouse_mode = True

                warehouse_groups = _detect_warehouse_groups(row, price_header_row, anchor_col + 1)
                if not warehouse_groups:
                    i += 1
                    continue

                delivery_time_col = -1
                for col_idx, value in enumerate(row):
                    cell_text = str(value).strip() if value is not None else ""
                    if "时效" in cell_text or "开始计" in cell_text:
                        delivery_time_col = col_idx
                        break

                if region_col == -1:
                    region_col = anchor_col + 1

                i += 2
                while i < len(all_rows):
                    data_row = all_rows[i]
                    region_val = ""
                    current_wh_codes: list[str] = []

                    if is_warehouse_mode:
                        region_val = (
                            str(data_row[anchor_col]).strip()
                            if anchor_col < len(data_row) and data_row[anchor_col] is not None
                            else ""
                        )
                        if region_val and region_val.upper() not in ["仓库代码", "仓库名称", "NONE", ""]:
                            current_wh_codes = extract_searchable_warehouse_codes(region_val)
                    else:
                        for check_col in range(max(0, region_col - 1), min(region_col + 3, len(data_row))):
                            candidate = str(data_row[check_col]).strip() if data_row[check_col] is not None else ""
                            if _is_region_row(candidate) or extract_searchable_warehouse_codes(candidate):
                                region_val = candidate
                                region_col = check_col
                                break

                    if not region_val or (is_warehouse_mode and not current_wh_codes):
                        any_val = [cell for cell in data_row if cell is not None]
                        if not any_val:
                            i += 1
                            continue
                        row_text = " ".join(str(cell) for cell in data_row if cell)
                        if any(keyword in row_text for keyword in ["赔付", "注意事项", "重货优惠", "下单产品"]):
                            break
                        i += 1
                        continue

                    region_clean = re.sub(r"[\(（].*?[\)）]", "", region_val).strip()
                    target_codes = current_wh_codes if is_warehouse_mode else [""]

                    for group in warehouse_groups:
                        prices = {}
                        for col_idx, tier_label in group["headers"]:
                            value = data_row[col_idx] if col_idx < len(data_row) else None
                            price = _safe_float(value)
                            if price is not None and price > 0:
                                prices[tier_label] = price

                        if not prices:
                            continue

                        delivery_time = ""
                        if delivery_time_col != -1 and delivery_time_col < len(data_row):
                            delivery_time = str(data_row[delivery_time_col] or "").strip()

                        for code in target_codes:
                            all_quotes.append(
                                {
                                    "渠道": f"{current_channel}({group['name']})",
                                    "目的地区": region_clean,
                                    "仓库代码": code,
                                    "时效和赔偿约定": "",
                                    "价格体系": prices,
                                    "起收重量": "",
                                    "宣称时效": delivery_time,
                                    "附加备注": f"来源: {region_val}",
                                    "_source": source_name,
                                    "_type": "warehouse_based" if is_warehouse_mode else "region_based",
                                }
                            )
                    i += 1

    except Exception as exc:
        import traceback

        print(f"Error parse_jinlian_excel {file_path}: {exc}")
        traceback.print_exc()
    finally:
        if wb:
            wb.close()

    return all_quotes
