import re
import openpyxl

def parse_complex_excel(file_path: str) -> list:
    """
    解析复杂的代理商分块报价表（升级版）

    三个核心升级：
    补丁一：用 openpyxl data_only=True 读取公式的真实计算结果
    补丁二：识别"双仓模式"（如义乌仓 + 泉州仓/厦门仓），分组读取每组价格
    补丁三：每个 Sheet 内支持存在多个"仓库代码锚点块"，不再只读第一块
    """
    blacklist = ["目录", "船期", "赔偿说明", "发票模版", "全渠道查价", "偏远邮编",
                 "无服务", "美国亚马逊仓库", "美国海运偏远", "夏威夷", "收货地址",
                 "FBA卡派查价表"]

    def is_valid_sheet(name: str) -> bool:
        for b in blacklist:
            if b in name:
                return False
        return True

    def safe_float(val):
        """安全地将值转换为浮点数，如果不能转换则返回 None"""
        if val is None:
            return None
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val).strip()
        try:
            return float(s)
        except (ValueError, TypeError):
            return None

    all_quotes = []

    try:
        # 补丁一：使用 openpyxl data_only=True，读取公式的真实计算结果（而非公式字符串）
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

        for sheet_name in wb.sheetnames:
            if not is_valid_sheet(sheet_name):
                continue

            ws = wb[sheet_name]
            # 将所有行读入内存，方便随机访问
            all_rows = []
            for row in ws.iter_rows(values_only=True):
                all_rows.append(list(row))

            source_name = file_path.split("\\")[-1].split("/")[-1]

            # 补丁三：允许整个 Sheet 里有多个"仓库代码"锚点块
            i = 0
            current_channel_name = sheet_name  # 渠道名默认为 Sheet 名
            while i < len(all_rows):
                row = all_rows[i]

                # 识别渠道名行（首列是主要文字描述，且不为空，且行内大部分列为空）
                # 用于捕获类似"华东明星达14T卡派\n\n船司：美森..."这样的子渠道标题行
                first_val = str(row[0]).strip() if row[0] is not None else ""
                non_empty_count = sum(1 for c in row if c is not None)
                is_channel_name = (
                    first_val
                    and non_empty_count <= 3
                    and "仓库" not in first_val
                    and any(kw in first_val for kw in ["卡派", "海派", "普卡", "快船", "快线", "经济", "特惠", "海铁"])
                    and len(first_val.split("\n")[0].strip()) <= 30  # 子渠道名通常不超过30字
                )
                if is_channel_name:
                    current_channel_name = first_val.split("\n")[0].strip()

                # 找"仓库代码"或"亚马逊仓库代码"锚点列
                wh_col_idx = -1
                for col_idx, val in enumerate(row):
                    cell_str = str(val).strip() if val is not None else ""
                    if "仓库代码" in cell_str or "仓库名称" in cell_str or "亚马逊仓库代码" in cell_str:
                        wh_col_idx = col_idx
                        break

                if wh_col_idx == -1:
                    i += 1
                    continue

                # ------- 处理一个锚点块 -------
                # 读取时效、赔偿约定信息（锚点行往往有这些关键词）
                block_type = ""
                for val in row:
                    s = str(val).strip() if val is not None else ""
                    if "赔偿" in s or "时效" in s or "开始计算" in s:
                        block_type += s + " "

                # ------- 补丁二：识别单仓 or 双仓模式 -------
                # 在锚点行中查找仓库组名（如"义乌仓"、"泉州仓/厦门仓"）
                warehouse_groups = []  # 格式：[{"name": "义乌仓", "start_col": 2}, ...]
                for col_idx, val in enumerate(row):
                    if col_idx <= wh_col_idx:
                        continue
                    s = str(val).strip() if val is not None else ""
                    # 检测仓库组名，如"义乌仓"、"泉州/厦门仓"等
                    if "仓" in s and s not in ["仓库代码", "仓库名称"] and len(s) > 1:
                        warehouse_groups.append({"name": s, "start_col": col_idx, "headers": []})

                # 读下一行的价格档位标题（如 12KG+、50KG+、350KG+、1CBM+）
                if i + 1 >= len(all_rows):
                    i += 1
                    continue

                header_row = all_rows[i + 1]

                if warehouse_groups:
                    # 双仓（或多仓）模式：为每个仓库组分配对应的价格列
                    for g_idx, group in enumerate(warehouse_groups):
                        next_group_start = warehouse_groups[g_idx + 1]["start_col"] if g_idx + 1 < len(warehouse_groups) else len(header_row)
                        for col_idx in range(group["start_col"], min(next_group_start, len(header_row))):
                            h = str(header_row[col_idx]).strip() if header_row[col_idx] is not None else ""
                            if h:
                                group["headers"].append((col_idx, h))
                else:
                    # 单仓模式（老逻辑兼容）：直接读锚点行的下一行
                    dummy_group = {"name": "", "start_col": wh_col_idx + 1, "headers": []}
                    for col_idx in range(wh_col_idx + 1, min(wh_col_idx + 12, len(header_row))):
                        h = str(header_row[col_idx]).strip() if header_row[col_idx] is not None else ""
                        stop_keywords = ["起收重量", "时效", "赔付", "备注"]
                        # 检查当前锚点行该列是否也有阻断关键词
                        anchor_val = str(row[col_idx]).strip() if col_idx < len(row) and row[col_idx] is not None else ""
                        if any(kw in anchor_val for kw in stop_keywords):
                            break
                        if h:
                            dummy_group["headers"].append((col_idx, h))
                    warehouse_groups = [dummy_group]

                # 找时效列、起收重量列、备注列的索引
                delivery_time_col = -1
                min_weight_col = -1
                notes_col = -1
                for col_idx, val in enumerate(row):
                    s = str(val).strip() if val is not None else ""
                    if "时效" in s or "开始计算" in s:
                        delivery_time_col = col_idx
                    if "起收重量" in s:
                        min_weight_col = col_idx
                    if "备注" in s:
                        notes_col = col_idx

                # 从第 i+2 行开始读数据行（锚点行 + 表头行 各占一行）
                i += 2
                while i < all_rows.__len__():
                    data_row = all_rows[i]

                    # 仓库代码列的值
                    wh_raw = data_row[wh_col_idx] if wh_col_idx < len(data_row) else None
                    wh_codes = str(wh_raw).strip() if wh_raw is not None else ""

                    # 遇到新锚点行或特殊中断行，退出当前块
                    if "仓库代码" in wh_codes or "仓库名称" in wh_codes:
                        break
                    if any(kw in wh_codes for kw in ["赔偿", "查验", "延误", "说明", "关税变动"]):
                        break
                    # 遇到全空行也退出
                    if all(c is None for c in data_row):
                        i += 1
                        continue

                    # 对每个仓库组，读取价格并生成一条记录
                    for group in warehouse_groups:
                        prices = {}
                        has_price = False
                        for col_idx, head in group["headers"]:
                            val = data_row[col_idx] if col_idx < len(data_row) else None
                            fval = safe_float(val)
                            if fval is not None and fval > 0:
                                prices[head] = fval
                                has_price = True

                        if not has_price:
                            continue

                        if not wh_codes:
                            continue

                        delivery_time = str(data_row[delivery_time_col]).strip() if delivery_time_col != -1 and delivery_time_col < len(data_row) and data_row[delivery_time_col] is not None else ""
                        min_weight = str(data_row[min_weight_col]).strip() if min_weight_col != -1 and min_weight_col < len(data_row) and data_row[min_weight_col] is not None else ""
                        notes = str(data_row[notes_col]).strip() if notes_col != -1 and notes_col < len(data_row) and data_row[notes_col] is not None else ""

                        # 预处理：如果整个单元格是"城市名(YYZ1-YYZ9/YDC5...)"格式，
                        # 先把括号内容提取出来，再按/分割
                        pre_wh = wh_codes
                        if re.search(r'[\u4e00-\u9fff]', pre_wh) and '(' in pre_wh:
                            inner = re.search(r'\(([^)]+)\)', pre_wh)
                            if inner:
                                pre_wh = inner.group(1)

                        # 拆分仓库代码（支持 /,，\n 空格等分隔符）
                        wh_codes_split = re.split(r'[/,，\n ]+', pre_wh)
                        cleaned_wh_list = []
                        for wh in wh_codes_split:
                            wh = wh.strip()
                            if not wh:
                                continue
                            # 排除含有中文字符的片段（城市名如"多伦多"）
                            if re.search(r'[\u4e00-\u9fff]', wh):
                                continue
                            # 去掉开头括号及之后（如"多伦多(YYZ1"里的城市名部分）
                            if re.match(r'[a-zA-Z\u4e00-\u9fff].*\(', wh):
                                wh = re.sub(r'^[^\(]*\(', '', wh)
                            # 去掉前后括号
                            wh = wh.strip('()（）')
                            # 处理形如 YYZ1-YYZ9 的范围描述，展开成逐个代码
                            range_match = re.match(r'^([A-Z]{3})(\d+)-([A-Z]{3})(\d+)$', wh.upper())
                            if range_match:
                                prefix, start, _, end = range_match.groups()
                                for n in range(int(start), int(end) + 1):
                                    cleaned_wh_list.append(f"{prefix}{n}")
                                continue
                            wh = wh.upper().strip('()（）')
                            # 最终过滤：必须是字母+数字组合的标准仓库代码，不含特殊字符
                            if re.match(r'^[A-Z]{2,4}\d+[A-Z]?$', wh):
                                cleaned_wh_list.append(wh)

                        for wh in cleaned_wh_list:

                            channel = current_channel_name
                            if group["name"]:
                                channel = f"{current_channel_name}({group['name']})"

                            all_quotes.append({
                                "渠道": channel,
                                "仓库代码": wh,
                                "时效和赔偿约定": block_type.strip(),
                                "价格体系": prices,
                                "起收重量": min_weight,
                                "宣称时效": delivery_time,
                                "附加备注": notes,
                                "_source": source_name
                            })
                    i += 1
                # 继续扫描下一个锚点块

    except Exception as e:
        import traceback
        print(f"Error parse_complex_excel {file_path}: {e}")
        traceback.print_exc()

    return all_quotes
