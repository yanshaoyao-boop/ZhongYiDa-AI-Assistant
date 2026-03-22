import re
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter


FIXED_FIELD_RULES = (
    {"source_header": "收货件数", "source_index": 29, "target_cell": "B16"},
    {"source_header": "额外服务", "source_index": 34, "target_cell": "F6"},
)


def load_a_headers(a_path):
    workbook = load_workbook(a_path, data_only=False, read_only=True)
    try:
        sheet = workbook.worksheets[0]
        return [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    finally:
        workbook.close()


def extract_template_targets(b_path):
    workbook = load_workbook(b_path, data_only=False, read_only=True)
    try:
        sheet = workbook.worksheets[0]
        targets = []

        for row in range(1, min(sheet.max_row, 16) + 1):
            left_label = sheet[f"A{row}"].value
            right_label = sheet[f"E{row}"].value
            if left_label:
                targets.append({"label": str(left_label).strip(), "cell": f"B{row}"})
            if right_label:
                targets.append({"label": str(right_label).strip(), "cell": f"F{row}"})

        for column in range(2, sheet.max_column + 1):
            label = sheet.cell(17, column).value
            if label:
                # read_only 模式下 EmptyCell 没有 .coordinate，用 get_column_letter 手动构建
                col_letter = get_column_letter(column)
                targets.append({"label": str(label).strip(), "cell": f"{col_letter}18"})

        return targets
    finally:
        workbook.close()


def generate_workbooks(a_path, b_path, mappings, output_dir, image_config=None):
    a_path = Path(a_path)
    b_path = Path(b_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    source_workbook = load_workbook(a_path, data_only=False, read_only=True)
    created_files = []

    try:
        source_sheet = source_workbook.worksheets[0]
        headers = load_a_headers(a_path)
        rows = list(source_sheet.iter_rows(min_row=2, values_only=True))

        for row in rows:
            if not any(value is not None and value != "" for value in row):
                continue

            filename_source = _get_value(row, headers, "客户单号", 6)
            if filename_source in (None, ""):
                continue

            workbook = load_workbook(b_path, data_only=False)
            try:
                sheet = workbook.worksheets[0]

                for mapping in mappings:
                    value = _get_value(row, headers, mapping["source_header"], None)
                    sheet[mapping["target_cell"]] = value

                for rule in FIXED_FIELD_RULES:
                    value = _get_value(row, headers, rule["source_header"], rule["source_index"])
                    sheet[rule["target_cell"]] = value

                fba_code = _get_value(row, headers, "客户单号", 6)
                piece_count = _get_value(row, headers, "收货件数", 29)
                if fba_code not in (None, "") and piece_count not in (None, ""):
                    # 判断件数是否为 1
                    try:
                        count_int = int(float(piece_count)) if piece_count is not None else 0
                    except (ValueError, TypeError):
                        count_int = 0
                    
                    if count_int == 1:
                        sheet["A18"] = f"{fba_code}U000001"
                    else:
                        sheet["A18"] = f"{fba_code}U000001-{piece_count}"

                if image_config:
                    img = Image(image_config["path"])
                    sheet.add_image(img, image_config["target_cell"])

                safe_name = sanitize_filename(str(filename_source))
                file_path = output_dir / f"{safe_name}.xlsx"
                workbook.save(file_path)
                created_files.append(file_path)
            finally:
                workbook.close()
    finally:
        source_workbook.close()

    zip_path = output_dir.parent / f"{output_dir.name}.zip"
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for file_path in created_files:
            archive.write(file_path, arcname=file_path.name)

    return {"count": len(created_files), "files": [str(path) for path in created_files], "zip_path": str(zip_path)}


def sanitize_filename(value):
    cleaned = re.sub(r'[<>:"/\\|?*]+', "_", value).strip()
    return cleaned or "未命名"


def _get_value(row, headers, header_name, fallback_index):
    if header_name in headers:
        index = headers.index(header_name)
        if index < len(row):
            return row[index]
    if fallback_index is not None:
        index = fallback_index - 1
        if index < len(row):
            return row[index]
    return None
